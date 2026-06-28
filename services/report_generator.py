import pandas as pd
import io
import logging
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from typing import Dict, Any

logger = logging.getLogger(__name__)

class ReportGenerator:
    """Generates professional, multi-sheet Excel reports with formatting and charts."""
    
    def __init__(self, master_df: pd.DataFrame, data: Dict[str, pd.DataFrame], 
                 dq_report: pd.DataFrame, unmatched_report: pd.DataFrame, 
                 action_items: pd.DataFrame, ai_state: Dict[str, Any]):
        self.master_df = master_df
        self.data = data
        self.dq_report = dq_report
        self.unmatched_report = unmatched_report
        self.action_items = action_items
        self.ai_state = ai_state

    def generate_excel_bytes(self) -> bytes:
        logger.info("Generating comprehensive Excel report...")
        output = io.BytesIO()
        
        # Use openpyxl engine to allow rich formatting
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            
            # 1. Leadership Summary & Communications (Text-heavy)
            self._write_text_sheet(writer, "Leadership Summary", self.ai_state.get('final_report', 'No report available.'))
            self._write_text_sheet(writer, "Draft Communications", "\n\n---\n\n".join(self.ai_state.get('communications', [])))
            
            # 2. Escalation Summary (Filtered from Action Items)
            escalations = self.action_items[self.action_items['Severity'].isin(['High', 'Critical'])] if not self.action_items.empty else pd.DataFrame()
            if not escalations.empty:
                escalations.to_excel(writer, sheet_name="Escalation Summary", index=False)
            
            # 3. Owner Action Plan
            if not self.action_items.empty:
                owner_plan = self.action_items.sort_values(by=['Owner', 'Severity'])
                owner_plan.to_excel(writer, sheet_name="Owner Action Plan", index=False)
            
            # 4. Missing Data & Quality Report
            if not self.dq_report.empty:
                self.dq_report.to_excel(writer, sheet_name="Missing Data Report", index=False)
            if not self.unmatched_report.empty:
                self.unmatched_report.to_excel(writer, sheet_name="Unmatched Records", index=False)
                
            # 5. Cash Flow & Progress Report (Synthesized Data)
            self._generate_cash_flow_sheet(writer)
            self._generate_progress_sheet(writer)
            
            # Apply formatting across all sheets
            workbook = writer.book
            self._apply_global_formatting(workbook)
            
        logger.info("Excel report generated successfully.")
        return output.getvalue()

    def _write_text_sheet(self, writer, sheet_name: str, text: str):
        df = pd.DataFrame([text], columns=['Content'])
        df.to_excel(writer, sheet_name=sheet_name, index=False)

    def _generate_cash_flow_sheet(self, writer):
        # Basic synthesis for cash flow
        col_df = self.data.get('collections', pd.DataFrame())
        const_df = self.data.get('construction', pd.DataFrame())
        
        inflow = pd.to_numeric(col_df['Amount Collected'], errors='coerce').sum() if not col_df.empty and 'Amount Collected' in col_df.columns else 0
        outflow = pd.to_numeric(const_df['Cost Impact'], errors='coerce').sum() if not const_df.empty and 'Cost Impact' in const_df.columns else 0
        
        cf_data = {
            "Metric": ["Total Collections (Inflow)", "Total Construction Cost (Outflow)", "Net Cash Flow"],
            "Amount": [inflow, outflow, inflow - outflow]
        }
        df = pd.DataFrame(cf_data)
        df.to_excel(writer, sheet_name="Cash Flow Report", index=False)

    def _generate_progress_sheet(self, writer):
        # Basic progress summary by Project
        if not self.master_df.empty and 'Project' in self.master_df.columns:
            # Try to aggregate some metrics if they exist
            agg_dict = {}
            if 'Booking Value' in self.master_df.columns: agg_dict['Booking Value'] = 'sum'
            if 'Amount Collected' in self.master_df.columns: agg_dict['Amount Collected'] = 'sum'
            
            if agg_dict:
                # Convert to numeric first safely
                for col in agg_dict.keys():
                    self.master_df[col] = pd.to_numeric(self.master_df[col], errors='coerce')
                
                prog = self.master_df.groupby('Project').agg(agg_dict).reset_index()
                prog.to_excel(writer, sheet_name="Progress Report", index=False)
            else:
                pd.DataFrame({"Message": ["No aggregateable columns found."]}).to_excel(writer, sheet_name="Progress Report", index=False)
        else:
            pd.DataFrame({"Message": ["Master data unavailable for progress."]}).to_excel(writer, sheet_name="Progress Report", index=False)

    def _apply_global_formatting(self, workbook):
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4F81BD")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            
            # Freeze top row
            worksheet.freeze_panes = "A2"
            
            # Format headers
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border
                
            # Auto-adjust column widths and apply borders to all cells
            for col in worksheet.columns:
                max_length = 0
                column_letter = col[0].column_letter
                for cell in col:
                    if cell.value:
                        # For text heavy sheets, set a fixed width and wrap text
                        if sheet_name in ["Leadership Summary", "Draft Communications"] and column_letter == 'A':
                            max_length = 100
                            cell.alignment = Alignment(wrap_text=True, vertical="top")
                        else:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                    cell.border = border
                
                if max_length > 0:
                    adjusted_width = min(max_length + 2, 100) # Cap width at 100
                    worksheet.column_dimensions[column_letter].width = adjusted_width

            # Conditional Formatting for Escalation Summary and Owner Action Plan
            if sheet_name in ["Escalation Summary", "Owner Action Plan"]:
                # Find the Severity column
                severity_col_idx = None
                for idx, cell in enumerate(worksheet[1]):
                    if cell.value == "Severity":
                        severity_col_idx = idx + 1
                        break
                
                if severity_col_idx:
                    red_fill = PatternFill("solid", fgColor="FF4B4B")
                    orange_fill = PatternFill("solid", fgColor="FFA421")
                    
                    for row in range(2, worksheet.max_row + 1):
                        cell = worksheet.cell(row=row, column=severity_col_idx)
                        if cell.value == "Critical":
                            cell.fill = red_fill
                            cell.font = Font(color="FFFFFF", bold=True)
                        elif cell.value == "High":
                            cell.fill = red_fill
                            cell.font = Font(color="FFFFFF")
                        elif cell.value == "Medium":
                            cell.fill = orange_fill
                        
            # Add a chart to Cash Flow Report
            if sheet_name == "Cash Flow Report" and worksheet.max_row > 1:
                chart = BarChart()
                chart.title = "Cash Flow Overview"
                chart.y_axis.title = "Amount"
                chart.x_axis.title = "Metric"
                
                data = Reference(worksheet, min_col=2, min_row=1, max_col=2, max_row=worksheet.max_row)
                cats = Reference(worksheet, min_col=1, min_row=2, max_row=worksheet.max_row)
                
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(cats)
                
                worksheet.add_chart(chart, "D2")
